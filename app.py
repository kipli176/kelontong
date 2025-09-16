# ============================================
# APP.PY - Kasir Offline/Online dengan Flask
# ============================================

from io import BytesIO
from datetime import datetime, timedelta

from flask import (
    Flask, render_template, request, jsonify,
    g, send_file, redirect, url_for, flash, session
)
import psycopg2
from psycopg2 import pool
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from werkzeug.security import check_password_hash, generate_password_hash

# ============================================
# KONFIGURASI APLIKASI & DATABASE
# ============================================

app = Flask(__name__)
app.secret_key = "ganti_dengan_secret_random"

DB_CONFIG = {
    "host": "192.168.1.17",
    "port": 15432,
    # "host": "postgres",
    # "port": 5432,
    "dbname": "iin",
    "user": "kipli_user",
    "password": "kipli_password"
}

# Gunakan connection pool agar efisien
db_pool = psycopg2.pool.SimpleConnectionPool(
    minconn=1,
    maxconn=20,
    **DB_CONFIG
)


# ============================================
# HELPER DATABASE & UTIL
# ============================================

def get_db():
    """Ambil koneksi database dari pool (per-request)."""
    if "db_conn" not in g:
        g.db_conn = db_pool.getconn()
    return g.db_conn


@app.teardown_appcontext
def close_db(exception=None):
    """Kembalikan koneksi ke pool setelah request selesai."""
    db_conn = g.pop("db_conn", None)
    if db_conn is not None:
        db_pool.putconn(db_conn)


def _parse_date(s, default):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return default

HARI_ID = {
    "Monday": "Senin",
    "Tuesday": "Selasa",
    "Wednesday": "Rabu",
    "Thursday": "Kamis",
    "Friday": "Jumat",
    "Saturday": "Sabtu",
    "Sunday": "Minggu",
}

# ============================================
# HELPER QUERY INTERNAL
# ============================================


def get_current_user():
    """Ambil user yang sedang login dari session."""
    if "user_id" not in session:
        return None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT u.id, u.nama, u.username, u.role,
               t.id, t.nama, t.kode, t.alamat
        FROM users u
        JOIN toko t ON u.toko_id = t.id
        WHERE u.id = %s
    """, (session["user_id"],))
    row = cur.fetchone()
    cur.close()
    if row:
        return {
            "id": row[0],
            "nama": row[1],
            "username": row[2],
            "role": row[3],
            "toko": {
                "id": row[4],
                "nama": row[5],
                "kode": row[6],
                "alamat": row[7],
            }
        }
    return None

def login_required(fn):
    """Decorator sederhana untuk proteksi route."""
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

# ============================================
# ROUTE HANDLER (HTML VIEW)
# ============================================

@app.route("/")
@login_required
def kasir():
    """Halaman utama kasir."""
    user = get_current_user()   # sudah ada di app.py
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, nama, no_hp FROM pembeli ORDER BY nama")
    pembeli = cur.fetchall()
    cur.close()
    return render_template("kasir.html", pembeli=pembeli, toko=user["toko"])

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, password_hash FROM users WHERE username=%s", (username,))
        row = cur.fetchone()
        cur.close()

        if row and check_password_hash(row[1], password):
            session["user_id"] = row[0]
            return redirect(url_for("kasir"))
        else:
            return render_template("login.html", error="Username atau password salah")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        nama_toko = request.form["nama_toko"].strip()
        kode_toko = request.form["kode_toko"].strip().lower()
        alamat    = request.form["alamat"].strip()
        nama_user = request.form["nama_user"].strip()
        username  = request.form["username"].strip().lower()
        password  = request.form["password"].strip()

        if not (nama_toko and kode_toko and nama_user and username and password):
            return render_template("register.html", error="Semua field wajib diisi")

        conn = get_db()
        cur = conn.cursor()
        try:
            # buat toko baru
            cur.execute("""
                INSERT INTO toko (nama, kode, alamat)
                VALUES (%s, %s, %s)
                RETURNING id
            """, (nama_toko, kode_toko, alamat))
            toko_id = cur.fetchone()[0]

            # buat user baru
            cur.execute("""
                INSERT INTO users (nama, username, password_hash, role, toko_id)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (
                nama_user,
                username,
                generate_password_hash(password),
                "admin",
                toko_id
            ))
            user_id = cur.fetchone()[0]

            conn.commit()

            # auto login setelah register
            session["user_id"] = user_id
            return redirect(url_for("kasir"))
        except Exception as e:
            conn.rollback()
            return render_template("register.html", error=f"Gagal register: {e}")
        finally:
            cur.close()

    return render_template("register.html")

@app.route("/penjualan")
@login_required
def penjualan():
    user = get_current_user()
    today = datetime.now().date()

    # ambil parameter filter dari query string
    d1 = _parse_date(request.args.get("start", ""), today)
    d2 = _parse_date(request.args.get("end", ""), today)

    conn = get_db()
    cur = conn.cursor()

    # transaksi per nota
    cur.execute("""
        SELECT id, tanggal, tx8, nama, no_hp, metode_bayar, total, laba, jml_item
        FROM v_penjualan_hari_ini
        WHERE DATE(tanggal) BETWEEN %s AND %s
          AND toko_id = %s
        ORDER BY tanggal DESC
    """, (d1, d2, user["toko"]["id"]))
    rows = cur.fetchall()

    # rekap barang
    cur.execute("""
        SELECT barcode, item_nama, harga_beli, harga_jual, total_qty, total_penjualan, total_laba
        FROM v_penjualan_rekap_barang_hari_ini
        WHERE toko_id = %s
          AND tgl BETWEEN %s AND %s
        ORDER BY item_nama, harga_jual
    """, (user["toko"]["id"], d1, d2))
    detail_rows = cur.fetchall()

    cur.close()

    # format tanggal untuk judul
    if d1 == d2:
        hari = HARI_ID[d1.strftime("%A")]
        keterangan_tanggal = f"{hari}, {d1.strftime('%d %B %Y')}"
    else:
        hari1 = HARI_ID[d1.strftime("%A")]
        hari2 = HARI_ID[d2.strftime("%A")]
        keterangan_tanggal = f"{hari1}, {d1.strftime('%d %B %Y')} s/d {hari2}, {d2.strftime('%d %B %Y')}"

    return render_template(
        "penjualan_hari_ini.html",
        rows=rows,
        detail_rows=detail_rows,
        toko=user["toko"],
        start=d1.strftime("%Y-%m-%d"),
        end=d2.strftime("%Y-%m-%d"),
        keterangan_tanggal=keterangan_tanggal
    )

@app.route("/penjualan-hari-ini/print-transaksi")
@login_required
def print_transaksi_hari_ini():
    user = get_current_user()
    today = datetime.now().date()
    d1 = _parse_date(request.args.get("start", ""), today)
    d2 = _parse_date(request.args.get("end", ""), today)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT tanggal, tx8, nama, no_hp, metode_bayar, jml_item, total, laba
        FROM v_penjualan_hari_ini
        WHERE DATE(tanggal) BETWEEN %s AND %s
          AND toko_id = %s
        ORDER BY tanggal DESC
    """, (d1, d2, user["toko"]["id"]))
    rows = cur.fetchall()
    cur.close()

    if d1 == d2:
        hari = HARI_ID[d1.strftime("%A")]
        keterangan_tanggal = f"{hari}, {d1.strftime('%d %B %Y')}"
    else:
        hari1 = HARI_ID[d1.strftime("%A")]
        hari2 = HARI_ID[d2.strftime("%A")]
        keterangan_tanggal = f"{hari1}, {d1.strftime('%d %B %Y')} s/d {hari2}, {d2.strftime('%d %B %Y')}"

    return render_template(
        "print_transaksi_hari_ini.html",
        rows=rows,
        toko=user["toko"],
        start=d1.strftime("%Y-%m-%d"),
        end=d2.strftime("%Y-%m-%d"),
        keterangan_tanggal=keterangan_tanggal
    )

@app.route("/penjualan-hari-ini/print-detail")
@login_required
def print_detail_hari_ini():
    user = get_current_user()
    today = datetime.now().date()
    d1 = _parse_date(request.args.get("start", ""), today)
    d2 = _parse_date(request.args.get("end", ""), today)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT barcode, item_nama, harga_beli, harga_jual, total_qty, total_penjualan, total_laba
        FROM v_penjualan_rekap_barang_hari_ini
        WHERE toko_id = %s
          AND tgl BETWEEN %s AND %s
        ORDER BY item_nama, harga_jual
    """, (user["toko"]["id"], d1, d2))
    rows = cur.fetchall()
    cur.close()

    if d1 == d2:
        hari = HARI_ID[d1.strftime("%A")]
        keterangan_tanggal = f"{hari}, {d1.strftime('%d %B %Y')}"
    else:
        hari1 = HARI_ID[d1.strftime("%A")]
        hari2 = HARI_ID[d2.strftime("%A")]
        keterangan_tanggal = f"{hari1}, {d1.strftime('%d %B %Y')} s/d {hari2}, {d2.strftime('%d %B %Y')}"

    return render_template(
        "print_detail_hari_ini.html",
        rows=rows,
        toko=user["toko"],
        start=d1.strftime("%Y-%m-%d"),
        end=d2.strftime("%Y-%m-%d"),
        keterangan_tanggal=keterangan_tanggal
    )
# ============================================
# API DATA (LAPORAN JSON)
# ============================================


@app.route("/api/detail-barang/<barcode>/<harga>")
@login_required
def api_detail_barang(barcode, harga):
    user = get_current_user()
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT LEFT(p.client_tx_id::text, 8) AS tx8,
               p.tanggal,
               COALESCE(pb.nama,'') AS pembeli,
               COALESCE(pb.no_hp,'') AS no_hp,
               d.qty
        FROM penjualan p
        JOIN penjualan_detail d ON d.penjualan_id = p.id
        LEFT JOIN pembeli pb ON pb.id = p.pembeli_id
        WHERE DATE(p.tanggal) = CURRENT_DATE
          AND p.toko_id = %s
          AND d.barcode = %s
          AND d.harga_jual = %s
        ORDER BY p.tanggal
    """, (user["toko"]["id"], barcode, harga))
    rows = cur.fetchall()
    cur.close()

    result = []
    for tx8, tanggal, pembeli, no_hp, qty in rows:
        result.append({
            "tx8": tx8,
            "waktu": tanggal.strftime("%H:%M:%S"),
            "pembeli": pembeli,
            "no_hp": no_hp,
            "qty": qty
        })
    return jsonify(result)


# ============================================
# EXPORT EXCEL
# ============================================

@app.route("/penjualan-hari-ini/export-transaksi/xlsx")
@login_required
def export_transaksi_hari_ini_xlsx():
    user = get_current_user()
    today = datetime.now().date()
    d1 = _parse_date(request.args.get("start", ""), today)
    d2 = _parse_date(request.args.get("end", ""), today)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT tanggal, tx8, nama, no_hp, metode_bayar, jml_item, total, laba
        FROM v_penjualan_hari_ini
        WHERE DATE(tanggal) BETWEEN %s AND %s
          AND toko_id = %s
        ORDER BY tanggal DESC
    """, (d1, d2, user["toko"]["id"]))
    rows = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaksi"

    # Judul sheet
    judul = (f"Laporan Transaksi "
             f"{d1.strftime('%d %b %Y')}" if d1 == d2
             else f"Laporan Transaksi {d1.strftime('%d %b %Y')} s/d {d2.strftime('%d %b %Y')}")
    ws.append([user["toko"]["nama"]])
    ws.append([judul])
    ws.append([])

    headers = ["Waktu", "No Transaksi", "Pembeli", "Metode", "Item", "Total", "Laba"]
    ws.append(headers)

    t_item, t_total, t_laba = 0, 0, 0
    for tgl, tx8, nama, hp, metode, jml_item, total, laba in rows:
        ws.append([
            tgl.strftime("%d-%m-%Y %H:%M:%S"),
            f"TX-{tx8.upper()}",
            f"{nama or ''} {(hp or '')}",
            metode,
            jml_item,
            float(total or 0),
            float(laba or 0),
        ])
        t_item += jml_item or 0
        t_total += float(total or 0)
        t_laba += float(laba or 0)

    ws.append([])
    ws.append(["TOTAL", "", "", "", t_item, t_total, t_laba])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"transaksi_{d1:%Y%m%d}_{d2:%Y%m%d}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/penjualan-hari-ini/export-detail/xlsx")
@login_required
def export_detail_hari_ini_xlsx():
    user = get_current_user()
    today = datetime.now().date()
    d1 = _parse_date(request.args.get("start", ""), today)
    d2 = _parse_date(request.args.get("end", ""), today)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT barcode, item_nama, harga_beli, harga_jual, total_qty, total_penjualan, total_laba
        FROM v_penjualan_rekap_barang_hari_ini
        WHERE toko_id = %s
          AND tgl BETWEEN %s AND %s
        ORDER BY item_nama, harga_jual
    """, (user["toko"]["id"], d1, d2))
    rows = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Barang"

    # Judul sheet
    judul = (f"Laporan Rekap Barang "
             f"{d1.strftime('%d %b %Y')}" if d1 == d2
             else f"Laporan Rekap Barang {d1.strftime('%d %b %Y')} s/d {d2.strftime('%d %b %Y')}")
    ws.append([user["toko"]["nama"]])
    ws.append([judul])
    ws.append([])

    headers = ["Barcode", "Nama Barang", "Harga Beli", "Harga Jual", "Qty", "Total Penjualan", "Total Laba"]
    ws.append(headers)

    gqty, gtotal, glaba = 0, 0, 0
    for barcode, item, hb, hj, qty, total, laba in rows:
        ws.append([
            barcode,
            item,
            float(hb or 0),
            float(hj or 0),
            int(qty or 0),
            float(total or 0),
            float(laba or 0),
        ])
        gqty += qty or 0
        gtotal += float(total or 0)
        glaba += float(laba or 0)

    ws.append([])
    ws.append(["", "TOTAL", "", "", gqty, gtotal, glaba])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"rekap_barang_{d1:%Y%m%d}_{d2:%Y%m%d}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================
# API SINKRONISASI
# ============================================

@app.route("/api/pembeli")
def api_pembeli():
    """
    Ambil daftar pembeli dari database.
    Return JSON: [{id, nama, no_hp}, ...]
    """
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nama, COALESCE(no_hp,'')
        FROM pembeli
        ORDER BY nama
    """)
    rows = cur.fetchall()
    cur.close()

    return jsonify([
        {"id": r[0], "nama": r[1], "no_hp": r[2]}
        for r in rows
    ])


@app.route("/api/sync-pembeli", methods=["POST"])
def sync_pembeli():
    """
    Simpan pembeli baru dari frontend offline.
    JSON: { "nama": "...", "no_hp": "...", "alamat": "..." }
    """
    data = request.get_json() or {}
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO pembeli (nama, no_hp, alamat)
            VALUES (%s, %s, %s)
            ON CONFLICT (no_hp) DO UPDATE
            SET nama = EXCLUDED.nama,
                alamat = EXCLUDED.alamat
            RETURNING id
        """, (data.get("nama"), data.get("no_hp"), data.get("alamat")))
        new_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({"status": "ok", "id": new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        cur.close()


@app.route("/api/sync-transaksi", methods=["POST"])
def sync_transaksi():
    """
    Simpan transaksi dari frontend offline.
    JSON:
    {
      "client_tx_id": "uuid-string",
      "tanggal_client": "2025-09-14T12:34:00",
      "pembeli": 1,
      "metode_bayar": "tunai",
      "bayar": 50000,
      "kembalian": 6500,
      "toko_id": 1,
      "items": [
        {"barcode":"123", "nama":"Gula", "qty":2,
         "harga_jual":15000, "harga_beli":12000, "potongan":0}
      ]
    }
    """
    data = request.get_json() or {}
    conn = get_db()
    cur = conn.cursor()
    try:
        # Cek apakah transaksi sudah ada (idempotent)
        cur.execute("SELECT id FROM penjualan WHERE client_tx_id=%s",
                    (data["client_tx_id"],))
        if cur.fetchone():
            return jsonify({"status": "duplicate", "msg": "Transaksi sudah ada"})

        # Insert header penjualan
        cur.execute("""
            INSERT INTO penjualan
            (client_tx_id, tanggal, pembeli_id,
             metode_bayar, bayar, kembalian, toko_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data["client_tx_id"],
            data.get("tanggal_client"),
            data.get("pembeli"),
            data.get("metode_bayar"),
            data.get("bayar"),
            data.get("kembalian"),
            data.get("toko_id")  # ✅ wajib isi toko_id
        ))
        penjualan_id = cur.fetchone()[0]

        # Insert detail barang
        for item in data.get("items", []):
            cur.execute("""
                INSERT INTO penjualan_detail
                (penjualan_id, barcode, nama, qty,
                 harga_jual, harga_beli, potongan)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                penjualan_id,
                item["barcode"],
                item["nama"],
                item["qty"],
                item["harga_jual"],
                item["harga_beli"],
                item.get("potongan", 0)
            ))

        conn.commit()
        return jsonify({"status": "ok", "id": penjualan_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "msg": str(e)}), 500
    finally:
        cur.close()


@app.route("/api/send-wa", methods=["POST"])
def api_send_wa():
    """
    Proxy untuk kirim WhatsApp via service eksternal.
    Body JSON: { "number": "62xxxxxxxxxx", "message": "..." }
    """
    import requests

    payload = request.get_json() or {}
    number = (payload.get("number") or "").strip()
    message = (payload.get("message") or "").strip()

    # Validasi sederhana
    if not number or not message:
        return jsonify({"status": "error", "msg": "number & message wajib"}), 400
    if not number.startswith("62") or not number.isdigit():
        return jsonify({"status": "error", "msg": "format nomor harus 62..."}), 400

    try:
        r = requests.post(
            "https://blast.sukipli.work/send-message",
            json={"number": number, "message": message},
            timeout=10,
        )
        # teruskan response dari server WA
        try:
            return jsonify(r.json()), r.status_code
        except ValueError:
            return r.text, r.status_code, {
                "Content-Type": r.headers.get("Content-Type", "text/plain")
            }
    except Exception as e:
        return jsonify({"status": "error", "msg": f"gagal kirim WA: {e}"}), 502


@app.route("/api/penjualan/<int:pid>")
def api_penjualan_detail(pid):
    """
    Ambil detail transaksi berdasarkan ID penjualan.
    Return JSON: {header:{...}, items:[...]}
    """
    conn = get_db()
    cur = conn.cursor()
    try:
        # Header penjualan
        cur.execute("""
            SELECT p.id, p.client_tx_id, p.tanggal, p.metode_bayar, 
                p.bayar, p.kembalian,
                COALESCE(pb.nama,''), COALESCE(pb.no_hp,'')
            FROM penjualan p
            LEFT JOIN pembeli pb ON pb.id = p.pembeli_id
            WHERE p.id = %s
        """, (pid,))
        h = cur.fetchone()
        if not h:
            return jsonify({"status": "error", "msg": "not found"}), 404

        header = {
            "id": h[0],
            "client_tx_id": str(h[1]),   # ✅ pakai UUID dari DB
            "tanggal": h[2].isoformat(),
            "metode_bayar": h[3],
            "bayar": float(h[4] or 0),
            "kembalian": float(h[5] or 0),
            "pembeli_nama": h[6],
            "no_hp": h[7],
        }


        # Detail item
        cur.execute("""
            SELECT nama, qty, harga_jual, harga_beli, potongan
            FROM penjualan_detail
            WHERE penjualan_id = %s
            ORDER BY id
        """, (pid,))
        items = [{
            "nama": r[0],
            "qty": int(r[1]),
            "harga_jual": float(r[2]),
            "harga_beli": float(r[3]),
            "potongan": float(r[4] or 0)
        } for r in cur.fetchall()]

        return jsonify({"header": header, "items": items})
    finally:
        cur.close()

@app.route("/api/all-barang")
def api_all_barang():
    """
    Ambil semua barang yang pernah terbeli (cache master barang).
    Return JSON: [{barcode, nama, harga_jual, harga_beli}, ...]
    """
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT barcode, nama, harga_jual, harga_beli
        FROM v_barang_terbeli
        ORDER BY nama
    """)
    rows = cur.fetchall()
    cur.close()

    return jsonify([
        {
            "barcode": r[0],
            "nama": r[1],
            "harga_jual": float(r[2] or 0),
            "harga_beli": float(r[3] or 0)
        }
        for r in rows
    ])


@app.route("/api/barang/<barcode>")
@login_required
def api_barang(barcode):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT barcode, nama, harga_beli, harga_jual, terakhir_dibeli
        FROM v_barang_terbeli
        WHERE barcode = %s
    """, (barcode,))
    row = cur.fetchone()
    cur.close()
    if row:
        return jsonify({
            "barcode": row[0],
            "nama": row[1],
            "harga_beli": float(row[2] or 0),
            "harga_jual": float(row[3] or 0),
            "terakhir_dibeli": row[4].isoformat() if row[4] else None
        })
    return jsonify(None)

# ============================================
# MAIN ENTRY POINT
# ============================================

if __name__ == "__main__":
    import os

    host = os.getenv("FLASK_HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT", 5000))
    debug = os.getenv("FLASK_DEBUG", "false").lower() in ("1", "true", "yes")

    app.run(host=host, port=port, debug=debug)
